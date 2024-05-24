# -*- coding: utf-8 -*-
"""
__created__ = 20220721
__authour__ = 60208787
__ver__ = 0.1.1
"""

import modules.classy as classy

import logging
import glob, yaml
from pathlib import Path
from tqdm import tqdm
import pandas as pd
import openpyxl



def makde_df_from_openpyxl(openpyxl_file, sheetname):
    try:
        values = pd.DataFrame(openpyxl_file[sheetname].values)
    except KeyError:
        print(f'{sheetname} sheet not found')
    else:   
        values = values.set_index(values.columns[0])
        values.columns = values.iloc[0]
        values = values[1:].loc[:,:'Value'].dropna()
        return values

def processing_MPC_folders(config,mpc_logger):
    
    # Log file is good for individual checks, but list here is faster for checking if files exist
    with open(Path(config['parent_path'],"logfile_mpc_processed.txt"), 'r') as f:
        loglist = f.read().splitlines()
    
    for machine in config['machine_paths']: 

        machine_name = machine.split('/')[1]
        raw_machine_results_folder = Path(config['root_results_path'],f"{config['number_in_results_path']} {machine_name}","MPC","Raw")

        #Search and sort for all folders under each machine path
        list_of_MPC_folders = sorted(
                                    [str(f.resolve()) for f in raw_machine_results_folder.iterdir()
                                    if f.is_dir() 
                                    if 'NDS-WKS-SN' in str(f)
                                    if str(f) not in loglist],
                                    reverse=True)
        
        # Only using this as indicator, not as tracking failed files
        failed_folders_count = 0 
        # with open(logfile, 'a') as f:
        for folder in tqdm(list_of_MPC_folders, smoothing=0.1):
            
            ### Try to generate an object for each folder...if fail (like no results csv),
            ### then add the folder to a failed count
            try:
                #Custom MPC module for these objects
                MPC_obj = classy.MPC_results(folder)
                mpc_results = MPC_obj.read_results()
            except FileNotFoundError:
                failed_folders_count += 1
                mylogger.error(f"{folder} has no results file")
            else:
                try:
                    MPC_obj.write_MPC_to_MyQAFolder(mpc_results,raw_machine_results_folder)
                    mpc_logger.info(folder)
                except:
                    mylogger.exception(f"{folder}")
            
        print(f"{failed_folders_count} failed in {len(list_of_MPC_folders)} from {machine_name} as no results CSV...check manually \n")


def processing_results_files(config,myQA_logger):

    with open(f"{config['parent_path']}/logfile_myQA_processed.txt", 'r') as f:
        loglist = f.read().splitlines()
    

    for machine in config['machines']:

        print(f"Processing {machine} myQA results")
        raw_machine_results_folder = Path(config['root_results_path'],f"{config['number_in_results_path']} {machine}","MPC","Raw")
        # There is a '{machine}' within results_folder_path variable, hence machine=machine to apply
        list_of_results_files = sorted(
                                    [str(x.resolve()) 
                                    for x in raw_machine_results_folder.glob("Results_*.xlsx") # extractin
                                    if str(x) not in loglist],
                                    reverse=True)
    
        # Read in template xltx for MyQA and reuse here
        template = openpyxl.load_workbook(f"{config['parent_path']}/Results/Template.xltx")

        for file in tqdm(list_of_results_files, smoothing=0.1):
            processing_results_file(file,template)
            myQA_logger.info(file)

def processing_results_file(file,template):

    # Assumes template is already prepped and in openpyxl format

    mylogger.info(file)
    file_to_write_to = file.replace('\Raw','') 
    try:    
        results_file = openpyxl.load_workbook(file)
    except:
        mylogger.exception(file)
    else:
        check_names = set(results_file.sheetnames)
        check_temp = set(template.sheetnames)

        template_sheets = []

        with pd.ExcelWriter(file_to_write_to) as writer:
            # For each sheet *actually* in the MPC results file
            for sheet in results_file.sheetnames:
                ref_date = results_file[results_file.sheetnames[0]].cell(2,2).value
                # Case handing of actual sheets names
                # Could do this in a ResultsFile object to keep main tidy, but eh
                if sheet == '6xMVkVEnhancedCouch' and '6xMVkV' in results_file.sheetnames:               
                    #Create new df from existing workbook using function to keep more tidy
                    template_df = makde_df_from_openpyxl(template, '6xMVkV')            
                    #As book is generator not list, need to force past the first item/val to extract vals in loop
                    try:
                        values_6x = makde_df_from_openpyxl(results_file, '6xMVkV')
                        #As book is generator not list, need to force past the first item/val to extract vals in loop
                        values_6xext = makde_df_from_openpyxl(results_file, '6xMVkVEnhancedCouch')
                    except AttributeError:
                        mylogger.error('Seems to not have generated values df...break')
                        break
                    else:
                        for item in template_df.index:
                            if 'Enhanced' in item:
                                try:
                                    template_df.loc[item] = values_6xext.loc[item]
                                except KeyError:
                                    mylogger.warning(f"{item} doesn't exist in sheet {sheet}")    
                            else:
                                try:
                                    template_df.loc[item] = values_6x.loc[item]
                                except KeyError:
                                    mylogger.warning(f"{item} doesn't exist in sheet {sheet}")   
                        check_names.add('6xMVkV')
                        template_df.to_excel(writer,sheet_name='6xMVkV')
                        # results_file.remove(results_file['6xMVkVEnhancedCouch'])
        

                    
                elif sheet == '6xMVkVEnhancedCouch' and '6xMVkV' not in results_file.sheetnames:
                
                    template_df = makde_df_from_openpyxl(template, '6xMVkV')

                    try:
                        values = makde_df_from_openpyxl(results_file, '6xMVkVEnhancedCouch')
                        
                    except AttributeError:
                        mylogger.error('Seems to not have generated values df...break')
                        break
                    else:
                        for item in template_df.index:
                                try:
                                    template_df.loc[item] = values.loc[item]
                                except KeyError:
                                    mylogger.warning(f"{item} doesn't exist in sheet {sheet}")
                        check_names.add('6xMVkV')
                        template_df.to_excel(writer,sheet_name="6xMVkV")

                    
                elif sheet in template.sheetnames:
                    template_df = makde_df_from_openpyxl(template, sheet)

                    try:
                        values = makde_df_from_openpyxl(results_file, sheet)
                        
                    except AttributeError:
                        mylogger.error('Seems to not have generated values df...break')

                        break
                    else:
                        for item in template_df.index:
                                try:
                                    template_df.loc[item] = values.loc[item]
                                except KeyError:
                                    mylogger.warning(f"{item} doesn't exist in sheet {sheet}")
                        template_df.to_excel(writer,sheet_name=sheet)
            else:
                mylogger.warning(f"{sheet} not in template")
                pass
            
            temp_not_in_result_file = check_temp-check_names

            for sheet in temp_not_in_result_file:
                template_df = makde_df_from_openpyxl(template, sheet)
                template_df.to_excel(writer,sheet_name=sheet)

    
def logging_handler(logger,file_name: str,opts: str):
    handler_dev = logging.FileHandler(file_name,mode=opts)
    #SEND THROUGH EVERYTHING ABOVE INFO HERE
    handler_dev.setLevel(logging.INFO)
    handler_dev.setFormatter(logging.Formatter('%(levelname)s - %(message)s'))
    logger.addHandler(handler_dev)
    #SET GLOBAL LEVEL TO INFO
    logger.setLevel(logging.INFO)
    return logger

def writing_handler(logger,file_name: str,opts: str):
    handler_dev = logging.FileHandler(file_name,mode=opts)
    #SEND THROUGH EVERYTHING ABOVE INFO HERE
    handler_dev.setLevel(logging.INFO)
    handler_dev.setFormatter(logging.Formatter('%(message)s'))
    logger.addHandler(handler_dev)
    #SET GLOBAL LEVEL TO INFO
    logger.setLevel(logging.INFO)
    return logger

            
if __name__ == '__main__':
        
    #Assumes config is in same folder as main script
    try:
        with open('config.yaml','r') as stream:
            config = yaml.safe_load(stream)
    except:
        pass

    # Set up logger at top level of script
    mpc_logger = logging.getLogger('mpc_logger')
    myQA_logger = logging.getLogger('myQA_logger')
    mylogger = logging.getLogger('mylogger')

    ##########
    # Handler two is for writing out error messages files
    mylogger = logging_handler(mylogger, f"{config['parent_path']}/dev.log",opts='w')

    # For writing out processed files
    mpc_logger = writing_handler(mpc_logger, f"{config['parent_path']}/logfile_mpc_processed.txt",opts='a')
    myQA_logger = writing_handler(myQA_logger, f"{config['parent_path']}/logfile_myQA_processed.txt",opts='a')



    print('\n Starting checks of MPC Folders \n')
    processing_MPC_folders(config, mpc_logger)
    
    print('\n Starting processing of data for myQA \n')
    processing_results_files(config, myQA_logger)